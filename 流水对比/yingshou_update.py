"""
根据liushui.xlsx中的发票号码和记账日期，到应收表里把对应发票号码的收款情况改成记账日期
"""

import sys, os
sys.path.insert(1, os.path.abspath("../../invoice_helper"))

import shutil

import openpyxl
from datetime import datetime

import traceback
import pandas as pd
from utils import box_warning

filename = 'F:\\工惠文化\\未收账款\\2022年应收账款.xlsx'
bak_filename = 'F:\\工惠文化\\未收账款\\备份\\2022年应收账款.xlsx'


def yingshou_update():
    df = pd.read_excel("liushui_duibi.xlsx", dtype=str)

    print(".", end="")
    code_date_map = {}
    for i, row in df.iterrows():
        date = row['记账日期']
        code = row['发票号码']
        if pd.isnull(date):
            box_warning("发票“%s”没有记账日期" % code)
            continue
        date = date.replace("00:00:00", "").strip()
        date = datetime.strptime(date, '%Y-%m-%d').strftime('%m.%d')
        code_date_map[code] = date

    print(".", end="")

    wb = openpyxl.load_workbook(filename)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        if sheet['B6'].value != '发票号码':
            box_warning("Sheet也“%s”A6单元格不是发票号码，请修正" % sheet_name)

        if sheet['W6'].value != '收款情况':
            box_warning("Sheet也“%s”W6单元格不是收款情况，请修正" % sheet_name)

        for i in range(1, sheet.max_row + 1):
            code = sheet['B%d' % i].value
            if code in code_date_map:
                try:
                    value = sheet['W%d' % i].value
                    sheet['W%d' % i].value = code_date_map[code]
                    print("将发票“%s”的收款情况由“%s”改为“%s”" % (code, value, code_date_map[code]))
                except:
                    traceback.print_exc()
                    box_warning("发票“%s”修改出错，所在Sheet为: %s, 单元格: W%d" % (code, sheet_name, i))

        print(".", end="")

    shutil.copy(filename, bak_filename.replace(".xlsx", "%s.xlsx" % datetime.now().strftime('%Y%m%d%H%M')))
    wb.save(filename)
    wb.close()

    print("处理完成")


if __name__ == '__main__':
    yingshou_update()
