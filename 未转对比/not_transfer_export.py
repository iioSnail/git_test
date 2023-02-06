import sys, os
sys.path.insert(1, os.path.abspath(".."))

import pandas as pd
from utils import columns_strip, fill_merge_cells, merge_cell_and_export, columns_clean, unmerge_cell
import openpyxl


def export_not_transfer_items():
    excel_name = unmerge_cell("yingshou.xlsx")
    excel = pd.read_excel(excel_name, sheet_name=None, dtype=str)

    result_df_list = []

    index_column = '发票号码'
    exclude_columns = ['商品名称', '规格', '单位', '数量', '单价', '金额', '税率', '税额', '税收分类编码', '备注']

    for key, df in excel.items():
        df.columns = list(df.iloc[4])
        df = columns_clean(df)
        if df.columns[0] != '发票代码':
            print("错误：sheet页“%s”第6行第一个单元格不是发票代码，请确保表头在第6行!" % key)

        df = df[5:]

        # 找出需要删除的列
        drop_idx = []
        for i, value in enumerate(list(df['发票代码'])):
            if "发票类别" in str(value):
                drop_idx.append(i + 5)

            if '发票代码' in str(value):
                drop_idx.append(i + 5)

        df = df.drop(drop_idx)

        # 填补缺失项（合并单元格）
        df = fill_merge_cells(df, index_column, exclude_columns)

        df = df[df['收款情况'] == '未转']

        result_df_list.append(df)

        print('“%s”处理成功.' % key)

    df = pd.concat(result_df_list)
    df['价税合计'] = df['价税合计'].astype(float)

    merge_cell_and_export(df, "weizhuan.xlsx", '发票号码', exclude_columns)


def merge_cells():
    offset = 2

    index_column = '发票号码'
    wb = openpyxl.load_workbook("weizhuan_temp.xlsx")
    sheet = wb[wb.sheetnames[0]]
    df = pd.read_excel("yingshou_final.xlsx", header=0)

    merge_columns = ['发票代码', '发票号码', '购方企业名称', '购方税号', '银行账号', '地址电话',
                     '开票日期', '商品编码版本号', '单据号', '价税合计', '联系人', '联系方式',
                     '接单人', '收款情况']

    for item in list(df[index_column].drop_duplicates()):
        index = df[df[index_column] == item].index
        if len(index) <= 1:
            continue

        start_row = index[0] + offset
        end_row = index[-1] + offset

        for i, column in enumerate(df.columns):
            if column not in merge_columns:
                continue

            sheet.merge_cells(start_row=start_row, end_row=end_row, start_column=i + 1, end_column=i + 1)

    wb.save('weizhuan_final.xlsx')
    os.remove("weizhuan_temp.xlsx")


if __name__ == '__main__':
    export_not_transfer_items()

    print("生成未转出完毕，文件：weizhuan.xlsx")
