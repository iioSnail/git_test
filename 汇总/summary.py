import os

import openpyxl
import pandas as pd


def summarize(df):
    df = df.drop(df[pd.isnull(df['商品名称'])].index)

    df['发票号码'] = df['发票号码'].ffill()
    df['金额'] = df['金额'].astype(float)
    df['税额'] = df['税额'].astype(float)
    df.insert(loc=len(df.columns), column='税收合计', value=0.)
    df.insert(loc=len(df.columns), column='联系人', value="")
    df.insert(loc=len(df.columns), column='联系方式', value="")
    df.insert(loc=len(df.columns), column='接单人', value="")
    df.insert(loc=len(df.columns), column='是否收款', value="未转")

    for item in list(df['发票号码'].drop_duplicates()):
        index = df[df['发票号码'] == item].index
        df.loc[index, '发票代码'] = df.loc[index, '发票代码'].ffill()
        df.loc[index, '购方企业名称'] = df.loc[index, '购方企业名称'].ffill()
        df.loc[index, '购方税号'] = df.loc[index, '购方税号'].ffill()
        df.loc[index, '银行账号'] = df.loc[index, '银行账号'].ffill()
        df.loc[index, '地址电话'] = df.loc[index, '地址电话'].ffill()
        df.loc[index, '开票日期'] = df.loc[index, '开票日期'].ffill()
        df.loc[index, '商品编码版本号'] = df.loc[index, '商品编码版本号'].ffill()
        df.loc[index, '单据号'] = df.loc[index, '单据号'].ffill()


    for item in list(df['发票号码'].drop_duplicates()):
        if len(df[df['发票号码'] == item]) == 1:
            index = df[df['发票号码'] == item].index
            df.loc[index, '税收合计'] = df.loc[index, '金额'] + df.loc[index, '税额']
        else:
            indexes = df[df['发票号码'] == item].index
            index = df[(df['发票号码'] == item) & (df['商品名称'] == '小计')].index
            df.loc[indexes, '税收合计'] = (df.loc[index, '金额'] + df.loc[index, '税额']).values[0]


    # todo 暂时不求冲正号码
    # 求负数合计的对应正数的发票号码
    # minus_df = df[df['税收合计'] < 0]
    # for item in minus_df['发票号码'].drop_duplicates():
    #     amount = minus_df[minus_df['发票号码'] == item]['税收合计'].drop_duplicates().values[0]
    #     codes = '&'.join(list(df[df['税收合计'] == -amount]['发票号码'].drop_duplicates()))
    #
    #     index = df[df['发票号码'] == item].index
    #     df.loc[index, '冲正号码'] = codes

    df.to_excel("huizong_temp.xlsx", index=False)    # 导出中间状态文件


    return df

def merge_cell(df):
    offset = 2

    wb = openpyxl.load_workbook("huizong_temp.xlsx")
    sheet = wb[wb.sheetnames[0]]

    for item in list(df['发票号码'].drop_duplicates()):
        index = df[df['发票号码'] == item].index
        if len(index) <= 1:
            continue

        start_row = index[0] + offset
        end_row = index[-1] + offset
        sheet.merge_cells(start_row=start_row, end_row=end_row, start_column=1, end_column=1)
        sheet.merge_cells(start_row=start_row, end_row=end_row, start_column=2, end_column=2)
        sheet.merge_cells(start_row=start_row, end_row=end_row, start_column=3, end_column=3)
        sheet.merge_cells(start_row=start_row, end_row=end_row, start_column=4, end_column=4)
        sheet.merge_cells(start_row=start_row, end_row=end_row, start_column=5, end_column=5)
        sheet.merge_cells(start_row=start_row, end_row=end_row, start_column=6, end_column=6)
        sheet.merge_cells(start_row=start_row, end_row=end_row, start_column=7, end_column=7)
        sheet.merge_cells(start_row=start_row, end_row=end_row, start_column=8, end_column=8)
        sheet.merge_cells(start_row=start_row, end_row=end_row, start_column=9, end_column=9)
        sheet.merge_cells(start_row=start_row, end_row=end_row, start_column=19, end_column=19)
        sheet.merge_cells(start_row=start_row, end_row=end_row, start_column=20, end_column=20)
        sheet.merge_cells(start_row=start_row, end_row=end_row, start_column=21, end_column=21)
        sheet.merge_cells(start_row=start_row, end_row=end_row, start_column=22, end_column=22)
        sheet.merge_cells(start_row=start_row, end_row=end_row, start_column=23, end_column=23)
        sheet.merge_cells(start_row=start_row, end_row=end_row, start_column=24, end_column=24)

    wb.save("huizong_final.xlsx")


if __name__ == '__main__':
    print("正在生成文件，请稍后...")
    df = pd.read_excel("huizong.xlsx", header=5, dtype=str)
    df = summarize(df)
    merge_cell(df)

    os.remove("huizong_temp.xlsx")
    print("生成完毕，文件为：huizong_final.xlsx")