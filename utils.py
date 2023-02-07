import time
import tkinter as tk
import tkinter.messagebox
from pathlib import Path

import pandas as pd
import openpyxl
import random
import os

def columns_strip(df):
    i = 1
    columns = []
    for item in df.columns:
        if pd.isnull(item):
            columns.append(' ' * i)
            i += 1
        else:
            columns.append(item.strip())
    df.columns = columns
    return df


def columns_clean(df):
    df = columns_strip(df)
    columns = []
    for item in df.columns:
        if item == '是否收款':
            item = '收款情况'
        if item == '电话':
            item = '联系方式'
        columns.append(item)

    df.columns = columns
    return df


def fill_merge_cells(df, index_column, exclude_columns=None):
    if exclude_columns is None:
        exclude_columns = []

    df[index_column] = df[index_column].ffill()
    # Remove columns that do not need to merge.
    columns = list(df.columns)
    columns.remove(index_column)
    for item in exclude_columns:
        if item in columns:
            columns.remove(item)

    for item in list(df[index_column].drop_duplicates()):
        index = df[df[index_column] == item].index

        for column in columns:
            df.loc[index, column] = df.loc[index, column].ffill()

    return df


def merge_cell_and_export(df, filename, index_column, exclude_columns=None):
    if exclude_columns is None:
        exclude_columns = []

    temp_filename = str(random.randint(0, 99999)) + "_temp.xlsx"
    df.to_excel(temp_filename, index=False)

    offset = 2

    df = df.reset_index(drop=True)

    wb = openpyxl.load_workbook(temp_filename)
    sheet = wb[wb.sheetnames[0]]

    for item in list(df[index_column].drop_duplicates()):
        index = df[df[index_column] == item].index
        if len(index) <= 1:
            continue

        start_row = index[0] + offset
        end_row = index[-1] + offset

        for i, column in enumerate(df.columns):
            if column in exclude_columns:
                continue

            sheet.merge_cells(start_row=start_row, end_row=end_row, start_column=i + 1, end_column=i + 1)

    wb.save(filename)
    wb.close()
    os.remove(temp_filename)


def fuzzy_compare_str(str1: str, str2: str):
    set1 = set(str1)
    set2 = set(str2)

    inter = set1.intersection(set2)

    if len(inter) / len(set1) >= 0.8 or len(inter) / len(set2) >= 0.8:
        return True

    return False


def str_to_num(s):
    s = str(s)
    try:
        if '%' in s:
            return float(s.replace("%", "")) / 100
        else:
            return float(s)
    except:
        return 0


def df_strip(df):
    for column in df.columns:
        if 'str' in dir(df[column]):
            df[column] = df[column].str.strip()
    return df


def box_warning(msg, exit_=True):
    top = tkinter.Tk()
    top.wm_attributes('-topmost', 1)
    top.withdraw()
    top.update()
    tk.messagebox.showwarning("提示", msg)
    top.destroy()
    if exit_:
        exit(0)


# 拆分所有的合并单元格，并赋予合并之前的值
def unmerge_and_fill_cells(worksheet):
    all_merged_cell_ranges = list(
        worksheet.merged_cells.ranges
    )

    for merged_cell_range in all_merged_cell_ranges:
        merged_cell = merged_cell_range.start_cell
        worksheet.unmerge_cells(range_string=merged_cell_range.coord)

        for row_index, col_index in merged_cell_range.cells:
            cell = worksheet.cell(row=row_index, column=col_index)
            cell.value = merged_cell.value


def unmerge_cell(filename):
    wb = openpyxl.load_workbook(filename)
    for sheet_name in wb.sheetnames:
        print('.', end='')
        sheet = wb[sheet_name]
        unmerge_and_fill_cells(sheet)
    filename = filename.replace(".xls", "_temp.xls")
    wb.save(filename)
    wb.close()

    # openpyxl保存之后，再用pandas读取会存在公式无法读取到的情况，使用下面方式就可以了
    # 原理=使用windows打开excel，然后另存为一下
    from win32com.client import Dispatch
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(str(Path(".").absolute() / filename))  # 这里必须填绝对路径
    xlBook.Save()
    xlBook.Close()

    return filename
